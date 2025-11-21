from pathlib import Path
from typing import Any, Dict, List
from dotenv import load_dotenv
import os
from openpyxl import load_workbook
from google import genai
from .invariants import get_invariants



def load_conjectures_from_excel(excel_path: str) -> List[Dict[str, Any]]:
    p = Path(excel_path)
    if not p.exists():
        raise SystemExit("No excel file found.")

    wb = load_workbook(excel_path, data_only=True)
    ws = wb["Conjectures"]

    header_row = 1
    col_conjecture_index = None

    for cell in ws[header_row]:
        if cell.value == "Conjecture":
            col_conjecture_index = cell.column

    if col_conjecture_index is None:
        raise ValueError("No 'Conjecture' column found.")

    conjectures = []

    for index, row in enumerate(range(header_row+1, ws.max_row+1), start=1):
        cell = ws.cell(row=row, column=col_conjecture_index)
        conjecture_text = cell.value

        function_code = get_gemini_response(conjecture_text, index)

        with open(f"conj_export_{index}.py", "w", encoding="utf-8") as f:
            f.write(f"# Conjecture: {conjecture_text}\n\n")
            f.write(function_code)

        exec(function_code, globals())

        function_name = f"conj_{index}"

        conjecture = {
            "ID": f"export_{index}",
            "conjecture": conjecture_text,
            "subclass": "",
            "score_function": globals()[function_name]
        }
        print(conjecture)
        conjectures.append(conjecture)
    return conjectures

def get_gemini_response(conjecture: str, index: int):
    load_dotenv()
    api_key = os.getenv("GOOGLE_API_KEY")
    invariants_list = get_invariants()

    client = genai.Client(api_key=api_key)

    config = genai.types.GenerateContentConfig(
        temperature=0.0,
    )

    response = client.models.generate_content(
        model="gemini-3-pro-preview",
        contents=f"""
        Tu es un expert Python chargé de traduire une conjecture mathématique en une fonction de scoring pour le domaine de la théorie des graphes.

        CONJECTURE : "{conjecture}"

        TA TÂCHE :
        Écris UNIQUEMENT le code d'une fonction Python nommée `conj_{index}` qui détecte si un graphe réfute cette conjecture.

        SIGNATURE OBLIGATOIRE :
        def conj_1(G, min_size, max_size):
            # ... ton code ...
            # return float

        LOGIQUE DE RETOUR (SCORE) :
        - Retourne `None` si le graphe n'est pas valide (ex: non connexe si demandé, hors taille, etc.).
        - Retourne un score < 0.0 (NÉGATIF) SI ET SEULEMENT SI le graphe est un CONTRE-EXEMPLE.
        - Retourne un score >= 0.0 si la conjecture est vraie pour ce graphe.
        - Formule standard pour inégalité A <= B : retourner `-(B - A)`. (Si A > B, le score devient négatif).

        RÈGLES TECHNIQUES :
        1. Imports : Fais tes imports À L'INTÉRIEUR de la fonction (ex: `import networkx as nx`, `import numpy as np`).
        2. Invariants : Tu peux importer `from .invariants import *`. Les fonctions disponibles sont : {invariants_list}.
        3. Graphes Spéciaux :
           - Si la conjecture parle de "Graphe G", utilise l'objet `G` passé en argument.
           - Si la conjecture parle de "K_n", "C_n", ou "Hypergraphe", IGNORE les arêtes de `G`. Utilise juste `n = G.number_of_nodes()` 
           pour construire ta structure mathématique virtuellement.
        4. Précision : Recopie les coefficients numériques (polynômes) EXACTEMENT sans arrondir. Vérifie les puissances (x^2 vs x^3).

        EXEMPLE DE SORTIE ATTENDUE (Ne pas inclure de markdown, juste le code) :
        def conj_1(G, min_size, max_size):
            import networkx as nx
            n = G.number_of_nodes()
            if n < min_size or n > max_size: return None
            # ... logique ...
            return score
        """,
        config=config
    )

    return extract_code(response.text)


def extract_code(content: str) -> str:
    lines = content.strip().splitlines()
    if not lines: return content.strip()
    # Nettoyage des balises ```python éventuelles
    if lines[0].strip().startswith("```"): lines = lines[1:]
    if lines and lines[-1].strip().startswith("```"): lines = lines[:-1]
    return "\n".join(lines).strip()