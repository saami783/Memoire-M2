from pathlib import Path
from openpyxl import Workbook, load_workbook

HEADERS = [
    "Id",
    "Titre de l'article",
    "Auteur",
    "DOI",
    "Date de publication",
    "Source",
    "Lien de l'article",
    "Conjecture",
]

def create_excel_file(excel_file: str, sheet_name: str = "Sheet1"):
    p = Path(excel_file)
    if p.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(HEADERS)
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_file)

def open_or_create_excel(excel_file: str, sheet_name: str = "Sheet1"):
    p = Path(excel_file)
    if not p.exists():
        create_excel_file(excel_file, sheet_name)
    wb = load_workbook(excel_file)
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(HEADERS)
    return wb

def get_next_excel_id(ws) -> int:
    """
    Retourne l'ID interne à utiliser pour la prochaine ligne.
    Cherche la dernière valeur non-vide de la colonne A et renvoie +1.
    """
    # Dernière ligne utilisée
    last_row = ws.max_row
    if last_row < 2:
        return 1  # rien d'écrit encore (hors en-têtes)
    last_val = ws.cell(row=last_row, column=1).value
    # Si la dernière ligne est vide (possible après suppressions), remonter jusqu’à trouver une valeur
    while last_row > 1 and (last_val is None or str(last_val).strip() == ""):
        last_row -= 1
        last_val = ws.cell(row=last_row, column=1).value
    try:
        return int(last_val) + 1 if last_val is not None else 1
    except Exception:
        # Si jamais quelqu’un a mis autre chose que du numérique en A, on repart à 1 + nb de lignes remplies
        return (last_row - 1) + 1  # nombre d'articles déjà écrits + 1
