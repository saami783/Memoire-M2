# todo : J'interroge un LLM pour qu'il me donne une fonction de score pour chaque conjecture de mon fichier xls, et je l'exécute dynamiquement
# en 1 : on ouvre le fichier excel, et pour chaque conjecture :
    # en 1 : Le LLM (température fixé pour la reproductibilité) génère une fonction de score sous forme de code Python pour une conjecture donnée.
    # en 2 : Le programme exécute dynamiquement ce code pour obtenir une fonction de score utilisable.
    # en 3 : Cette fonction de score est ensuite utilisée par le programme de Thibualt pour évaluer les graphes générés et mutés.
    # note : on doit logger plusieurs choses :
    # la fonction de score généré par le LLM, le temps mis pour trouvé le contre-exemple,

# ça se passera dans refutation_heuristics/local_search.py, dans la fonction _evaluate_graph.

"""
on va découper le code pour qu'il soit plus lisible.
Donc, comprend bien qu'il ne faut pas toucher au fonctionnement général du programme existant.
Lui ce qu'il fait c'est que s'il y a des ID dans le fichier identifiers.txt alors
il va chercher les conjectures du benchmark.csv, si ils existent alors il exécute une fonction de score générique,
sinon si l'id n'est pas dans le benchmark.csv mais que ça correspond à un ID d'une fonction de conjecture.
par exemple "a" alors il exécutera la fonction conj_a. Dans tous les cas, il exécute une fonction de score de la conjecture,
donc c'est ici qu'il faut que nous branchions notre logique, où, si dans identifiers.txt il y a marqué 'export' par exemple,
alors c'est ma logique à moi qui sera prise en compte en exécutant la fonction de score dynamique.
"""
from pathlib import Path
from typing import Any, Dict, List
from dotenv import load_dotenv
import os
from openpyxl import Workbook, load_workbook
from mistralai import Mistral

"""
# Supposons que le LLM génère le code suivant pour une conjecture donnée
function_code =
def conj_x(G, min_size, max_size):
    la logique 
    return le calcul de la fonction de score.

# Exécutee dynamiquement le code pour obtenir la fonction
exec(function_code, globals())

# Maintenant, la fonction conj_x est disponible dans le namespace global
score = conj_x(graph, min_size, max_size)

"""

def load_conjectures_from_excel(excel_path: str) -> List[Dict[str, Any]]:
    p = Path(excel_path)
    if not p.exists():
        raise SystemExit("No excel file found.")

    wb = load_workbook(excel_path, data_only=True)
    ws = wb["Conjectures"]

    header_row = 1
    col_conjecture_index = None

    for cell in ws[header_row]:
        if cell.value == "Conjecture":
            col_conjecture_index = cell.column

    if col_conjecture_index is None:
        raise ValueError("No 'Conjecture' column found.")

    conjectures = []

    for index, row in enumerate(range(header_row+1, ws.max_row+1), start=1):
        cell = ws.cell(row=row, column=col_conjecture_index)
        conjecture_text = cell.value

        function_code = generate_score_function_with_llm(conjecture_text, index)
        exec(function_code, globals())

        function_name = f"conj_{index}"
        conjecture = {
            "ID": f"export_{index}",
            "conjecture": conjecture_text,
            "subclass": "",
            "score_function": globals()[function_name]
        }
        print(conjecture)
        conjectures.append(conjecture)
        break
    return conjectures

def generate_score_function_with_llm(conjecture: str, index: int):
    # todo : tester les 3 modèles
    # model = "mistral-large-latest"
    # model = "ministral-8b-2410"
    model = "mistral-large-2411"
    return get_mistral_reponse(model, conjecture, index)

def get_mistral_reponse(model: str, conjecture: str, index: int) -> str:
    """
    La signature de la fonction doit obligatoirement être sous la forme suivante : conj_{index}(G, min_size, max_size).
    """
    load_dotenv()
    api_key = os.getenv("MISTRAIL_API_KEY_PRO")
    client = Mistral(api_key=api_key)
    chat = client.chat.complete(
        temperature=0,
        top_p=1,
        model=model,
        messages=[
            {
                "role": "system",
                "content": "Vous êtes un assistant spécialisé dans la génération de fonctions de score pour des conjectures en théorie des graphes."
            },
            {
                "role": "user",
                "content": f"""
                Voici une conjecture en théorie des graphes :
                {conjecture}

                Générez une fonction de score en Python qui évalue si un graphe est un contre-exemple à cette conjecture.
                La fonction doit avoir la signature suivante : conj_{index}(G, min_size, max_size).
                La fonction doit retourner None si le graphe n'est pas éligible (taille hors limites ou non conforme à la conjecture).
                Sinon, elle doit retourner un score numérique qui est négatif si le graphe est un contre-exemple à la conjecture.

                Exemples de fonctions de score :
                def conj_a(G, min_size, max_size):
                    order = G.number_of_nodes()
                    if order < min_size or order > max_size:
                        return None
                    if not binary_properties_functions["connected"](G):
                        return None
                    lambda_max = invariants_functions["largest_eigenvalue"](G)
                    matching_number = invariants_functions["matching_number"](G)
                    return - (math.sqrt(order - 1) + 1 - lambda_max - matching_number)

                def conj_b(G, min_size, max_size):
                    order = G.number_of_nodes()
                    if order < min_size or order > max_size:
                        return None
                    if not binary_properties_functions["tree"](G):
                        return None
                    diameter = invariants_functions["diameter"](G)
                    k = math.floor(2 * diameter / 3)
                    proximity = invariants_functions["proximity"](G)
                    kth_largest_distance_eigenvalue = invariants_functions["kth_largest_distance_eigenvalue"](G, k)
                    return -(-proximity - kth_largest_distance_eigenvalue)
                
                def conj_c(G, min_size, max_size):
                    order = G.number_of_nodes()
                    if order < min_size or order > max_size:
                        return None
                    if not binary_properties_functions["tree"](G):
                        return None
                    pA, pD = invariants_functions["pA"](G), invariants_functions["pD"](G)
                    m = invariants_functions["m"](G)
                    return -(abs(pA / m  - (1 - pD / order)) - 0.28)
                                
                def conj_d(G, min_size, max_size):
                    order = G.number_of_nodes()
                    if order < min_size or order > max_size:
                        return None
                    if not binary_properties_functions["connected"](G):
                        return None
                    A = nx.adjacency_matrix(G).todense()
                    eigenvalues = np.linalg.eigvals(A)
                    second_largest_eigenvalues = np.sort(eigenvalues)[-2]
                    harmonic_index = invariants_functions["harmonic_index"](G)
                    return -(second_largest_eigenvalues - harmonic_index)

                Assurez-vous que la fonction est correctement indentée.
                Ne fournissez que le code de la fonction, sans commentaires ni texte supplémentaire. Vraiment uniquement la fonction est rien d'autre.
                """
            }
        ]
    )
    raw = chat.choices[0].message.content
    function = extract_code(raw)
    print(function)
    return function

def extract_code(content: str) -> str:
    """
    Enlève les ```python / ``` autour du code si présents
    et renvoie uniquement le code exécutable.
    """
    lines = content.strip().splitlines()

    if not lines:
        return content.strip()

    first = lines[0].strip().lower()
    if first.startswith("```"):
        lines = lines[1:]

    if lines and lines[-1].strip().startswith("```"):
        lines = lines[:-1]

    return "\n".join(lines).strip()