from pathlib import Path
import json
from openpyxl import load_workbook

HDR_ID = "Id"
HDR_FICHIER = "Fichier"

def map_headers(ws):
    """Retourne un dict {header -> index de colonne (0-based)} basé sur la ligne 1."""
    headers = {}
    for j, cell in enumerate(ws[1], start=0):
        if cell.value:
            headers[str(cell.value).strip()] = j
    return headers

def ensure_conj_sheet(wb, conj_sheet: str):
    """(Ré)initialise la feuille Conjectures avec l'en-tête, sans créer de ligne vide."""
    if conj_sheet in wb.sheetnames:
        ws = wb[conj_sheet]
        # purge toutes les lignes sauf l'entête
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)   # supprime de la 2e ligne jusqu'à la fin
        ws.cell(row=1, column=1, value="Article_id")
        ws.cell(row=1, column=2, value="Conjecture")
        ws.cell(row=1, column=3, value="Page")
    else:
        ws = wb.create_sheet(conj_sheet)
        ws.append(["Article_id", "Conjecture", "Page"])
    return ws


def update_excel_with_conjectures(excel_file: str, articles_sheet: str, conj_sheet:str, json_dir: Path):

    if not Path(excel_file).exists():
        raise FileNotFoundError(f"Fichier Excel introuvable: {excel_file}")
    if not json_dir.exists():
        raise FileNotFoundError(f"Dossier des JSON introuvable: {json_dir}")

    wb = load_workbook(excel_file)
    if articles_sheet not in wb.sheetnames:
        raise ValueError(f'La feuille "{articles_sheet}" est introuvable dans {excel_file}')

    ws_articles = wb[articles_sheet]
    ws_conj = ensure_conj_sheet(wb, conj_sheet)

    # Mappe les colonnes de "Articles"
    headers = map_headers(ws_articles)
    for needed in (HDR_ID, HDR_FICHIER):
        if needed not in headers:
            raise ValueError(f'Colonne "{needed}" absente de la feuille "{articles_sheet}".')

    col_id = headers[HDR_ID]
    col_fichier = headers[HDR_FICHIER]

    ajoutees = 0
    manquantes = 0
    sans_fichier = 0

    # parcours des lignes d'articles
    for i in range(2, ws_articles.max_row + 1):
        article_id = ws_articles.cell(row=i, column=col_id + 1).value
        file_name = ws_articles.cell(row=i, column=col_fichier + 1).value

        if not file_name:
            sans_fichier += 1
            continue

        # le json a le même nom que le pdf + ".json"
        json_path = json_dir / f"{file_name}.json"
        if not json_path.exists():
            # Pas de JSON pour cet article alors on marque "aucune conjecture".
            ws_conj.append([article_id, "aucune conjecture (json manquant)", ""])
            ajoutees += 1
            manquantes += 1
            continue

        try:
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            # json illisible on note l’erreur comme "aucune conjecture"
            ws_conj.append([article_id, f"aucune conjecture (json illisible: {e})", ""])
            ajoutees += 1
            continue

        # Lecture des conjectures
        contains = (data.get("contains_conjecture") or "").strip().lower()
        conjectures = data.get("conjectures") or []

        if contains == "yes" and len(conjectures) > 0:
            for c in conjectures:
                text = c.get("text", "").strip()
                page = c.get("page", "")
                ws_conj.append([article_id, text, page])
                ajoutees += 1
        else:
            # aucun élément exploitable, on met une ligne "aucune conjecture"
            ws_conj.append([article_id, "aucune conjecture", ""])
            ajoutees += 1

    wb.save(excel_file)
    print(f"Conjectures ajoutees: {ajoutees} | JSON manquants: {manquantes} | Articles sans 'Fichier': {sans_fichier}")
