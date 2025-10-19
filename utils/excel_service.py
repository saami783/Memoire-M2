from pathlib import Path
from openpyxl import Workbook, load_workbook

HEADERS_SHEET_ARTICLES = [
    "Id",
    "Titre de l'article",
    "Auteur",
    "DOI",
    "Date de publication",
    "Source",
    "Lien de l'article",
    "Fichier"
]

HEADERS_SHEET_CONJECTURES = [
    "Article_id",
    "Conjecture",
    "Page",
]

def create_excel_file(excel_file: str, sheet_name: str = "Feuille 1", sheet_name2: str = "Conjectures"):
    """Crée un Excel s'il n'existe pas, avec deux feuilles et leurs en-têtes."""
    p = Path(excel_file)
    if p.exists():
        return
    wb = Workbook()
    # Feuille 1
    ws1 = wb.active
    ws1.title = sheet_name
    ws1.append(HEADERS_SHEET_ARTICLES)
    # Feuille 2
    ws2 = wb.create_sheet(title=sheet_name2)
    ws2.append(HEADERS_SHEET_CONJECTURES)

    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(excel_file)

def open_or_create_excel(excel_file: str, sheet_name: str = "Feuille 1", sheet_name2: str = "Conjectures"):
    """Ouvre le classeur ; s’il n’existe pas, le crée avec les 2 feuilles + en-têtes.
       Garantit que les deux feuilles existent (sinon les crée)."""
    p = Path(excel_file)
    if not p.exists():
        create_excel_file(excel_file, sheet_name, sheet_name2)
    wb = load_workbook(excel_file)

    if sheet_name not in wb.sheetnames:
        ws1 = wb.create_sheet(title=sheet_name)
        ws1.append(HEADERS_SHEET_ARTICLES)
    if sheet_name2 not in wb.sheetnames:
        ws2 = wb.create_sheet(title=sheet_name2)
        ws2.append(HEADERS_SHEET_CONJECTURES)

    return wb

def get_next_excel_id(ws) -> int:
    """
    Retourne l'ID interne à utiliser pour la prochaine ligne.
    Cherche la dernière valeur non-vide de la colonne A et renvoie +1.
    """
    # Dernière ligne utilisée
    last_row = ws.max_row
    if last_row < 2:
        return 1  # rien d'écrit encore (hors en-têtes)
    last_val = ws.cell(row=last_row, column=1).value
    # Si la dernière ligne est vide (possible après suppressions), remonter jusqu’à trouver une valeur
    while last_row > 1 and (last_val is None or str(last_val).strip() == ""):
        last_row -= 1
        last_val = ws.cell(row=last_row, column=1).value
    try:
        return int(last_val) + 1 if last_val is not None else 1
    except Exception:
        # Si jamais quelqu’un a mis autre chose que du numérique en A, on repart à 1 + nb de lignes remplies
        return (last_row - 1) + 1  # nombre d'articles déjà écrits + 1
