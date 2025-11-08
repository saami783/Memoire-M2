import json
from pathlib import Path

from openpyxl import load_workbook

EXCEL_PATH = Path("articles.xlsx")

JSON_DIRS = [
    Path("extraction/mistral/json"),
    Path("extraction/deepseek/json"),
]


def _trouver_colonnes_par_nom(feuille):
    """
    Retourne un dict {nom_colonne: index_colonne} basé sur la première ligne.
    """
    mapping = {}
    for cell in feuille[1]:
        if cell.value is not None:
            mapping[str(cell.value).strip()] = cell.column
    return mapping


def _trouver_fichier_json(nom_fichier_sans_json: str) -> Path | None:
    """
    À partir du contenu de la colonne 'Fichier' (sans .json),
    cherche le fichier correspondant dans JSON_DIRS.
    """
    nom_json = nom_fichier_sans_json + ".json"
    for base_dir in JSON_DIRS:
        candidate = base_dir / nom_json
        if candidate.exists():
            return candidate
    return None


def _charger_json(path: Path) -> dict | None:
    """
    Charge un JSON et renvoie le dict, ou None en cas de souci.
    """
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        print(f"[WARN] Impossible de lire le JSON {path} : {e}")
        return None


def update_excel_with_conjectures():
    """
    Méthode principale (sans paramètres) à appeler depuis ton main.

    - Lit la feuille 'Articles'
    - Pour chaque ligne, ouvre le JSON associé si présent
    - Met à jour :
        * Articles!Contient conjecture = 'Oui' / 'Non'
        * ajoute des lignes dans la feuille 'Conjectures'
        * ajoute des lignes dans la feuille 'Parametres'
    """
    if not EXCEL_PATH.exists():
        raise FileNotFoundError(f"Fichier Excel introuvable : {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH)

    try:
        ws_articles = wb["Articles"]
        ws_conjectures = wb["Conjectures"]
        ws_parametres = wb["Parametres"]
    except KeyError as e:
        raise KeyError(f"Feuille manquante dans le fichier Excel : {e}")

    cols_articles = _trouver_colonnes_par_nom(ws_articles)

    col_id = cols_articles.get("Id")
    col_fichier = cols_articles.get("Fichier")
    col_contient_conj = cols_articles.get("Contient conjecture")

    if col_id is None or col_fichier is None or col_contient_conj is None:
        raise ValueError(
            "Les colonnes 'Id', 'Fichier' ou 'Contient conjecture' "
            "sont introuvables dans la feuille 'Articles'."
        )

    # --- Préparation des données à insérer ---
    lignes_conjectures = []  # tuples (Article_id, Conjecture, Fichier)
    lignes_parametres = []   # tuples (Conjecture_Id, Parametre)

    # --- Parcours de la feuille Articles ---
    for row_idx in range(2, ws_articles.max_row + 1):
        article_id = ws_articles.cell(row=row_idx, column=col_id).value
        fichier_base = ws_articles.cell(row=row_idx, column=col_fichier).value

        if article_id is None or fichier_base is None:
            # Ligne vide ou article incomplet : on ignore
            continue

        fichier_base = str(fichier_base).strip()
        json_path = _trouver_fichier_json(fichier_base)

        if json_path is None:
            # Pas de JSON correspondant : on met "JSON introuvable" et on continue
            ws_articles.cell(row=row_idx, column=col_contient_conj).value = "JSON introuvable"
            print(f"[INFO] JSON introuvable pour l'article {article_id} : {fichier_base}")
            continue

        data = _charger_json(json_path)
        if not data:
            ws_articles.cell(row=row_idx, column=col_contient_conj).value = "Non"
            continue

        contains_conjecture = bool(data.get("contains_conjecture", False))
        conjectures = data.get("conjectures", []) or []

        # Mise à jour de la colonne "Contient conjecture"
        ws_articles.cell(row=row_idx, column=col_contient_conj).value = "Oui" if contains_conjecture else "Non"

        if not contains_conjecture or not conjectures:
            continue

        # Pour chaque conjecture du JSON
        for conj in conjectures:
            texte_conj = conj.get("conjecture", "").strip()
            label = str(conj.get("label", "")).strip()
            params = conj.get("parameters", []) or []

            if not texte_conj:
                continue  # conjecture vide

            lignes_conjectures.append(
                (
                    article_id,
                    texte_conj,
                    fichier_base
                )
            )

            conjecture_id = f"{article_id}-{label if label else 'no_label'}"

            # Ajout des paramètres
            for p in params:
                param_text = str(p).strip()
                if not param_text:
                    continue
                lignes_parametres.append(
                    (
                        conjecture_id,
                        param_text
                    )
                )

    cols_conj = _trouver_colonnes_par_nom(ws_conjectures)
    col_c_article_id = cols_conj.get("Article_id")
    col_c_conjecture = cols_conj.get("Conjecture")
    col_c_fichier = cols_conj.get("Fichier")

    if col_c_article_id is None or col_c_conjecture is None or col_c_fichier is None:
        raise ValueError(
            "Les colonnes 'Article_id', 'Conjecture' ou 'Fichier' "
            "sont introuvables dans la feuille 'Conjectures'."
        )

    cols_param = _trouver_colonnes_par_nom(ws_parametres)
    col_p_conj_id = cols_param.get("Conjecture_id")
    col_p_param = cols_param.get("Paremetre") or cols_param.get("Parametre")

    if col_p_conj_id is None or col_p_param is None:
        raise ValueError(
            "Les colonnes 'Conjecture_id' ou 'Paremetre' sont introuvables "
            "dans la feuille 'Parametres'."
        )

    next_row_conj = ws_conjectures.max_row + 1
    for article_id, texte_conj, fichier_base in lignes_conjectures:
        ws_conjectures.cell(row=next_row_conj, column=col_c_article_id).value = article_id
        ws_conjectures.cell(row=next_row_conj, column=col_c_conjecture).value = texte_conj
        ws_conjectures.cell(row=next_row_conj, column=col_c_fichier).value = fichier_base
        next_row_conj += 1

    next_row_param = ws_parametres.max_row + 1
    for conj_id, param_text in lignes_parametres:
        ws_parametres.cell(row=next_row_param, column=col_p_conj_id).value = conj_id
        ws_parametres.cell(row=next_row_param, column=col_p_param).value = param_text
        next_row_param += 1

    wb.save(EXCEL_PATH)
    print("[OK] Mise à jour du fichier Excel terminée (ajout en fin de tables).")